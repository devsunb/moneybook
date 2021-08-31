import os
from datetime import datetime

from openpyxl import load_workbook, Workbook

from moneybook.util import comma_str_to_int
import numpy as np


class Excel:
    def __init__(self, path):
        self.path = path
        self.wb = None
        self.ws = None

        if os.path.exists(path):
            self.open(path)
        else:
            self.new()

    def open(self, path):
        self.wb = load_workbook(path)
        self.ws = self.wb.active

    def new(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def save(self):
        self.wb.save(self.path)

    def select_ws(self, name):
        self.ws = self.wb[name]

    def create_ws(self, name, index=None):
        self.ws = self.wb.create_sheet(name, index)

    def get_value(self, row, col):
        return self.ws.cell(row, col).value

    def get_values(self, min_row, min_col, max_row, max_col, dtype=np.float64):
        values = np.empty((max_row - min_row + 1, max_col - min_col + 1), dtype=dtype)
        for y, row in enumerate(self.ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col)):
            for x, cell in enumerate(row):
                values[y, x] = cell.value
        return values

    # def get_values(self, min_row=None, max_row=None, min_col=None, max_col=None):
    #     return [[datetime.strptime(r[0].value, '%Y.%m.%d %H:%M:%S'), r[1].value, comma_str_to_int(r[2].value),
    #              comma_str_to_int(r[3].value), r[4].value, r[5].value, r[6].value]
    #             for r in self.ws.iter_rows(min_row, max_row, min_col, max_col)]

    def set_value(self, row, col, value):
        self.ws.cell(row, col).value = value

    def set_values(self, min_row, min_col, values):
        max_row, max_col = np.array([min_row - 1, min_col - 1]) + values.shape
        for y, row in enumerate(self.ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col)):
            for x, cell in enumerate(row):
                cell.value = values[y, x]
